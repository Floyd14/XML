# -*- coding: UTF-8 -*-
""" xml_To_xlsx Test1.
This program convert .xml file to .xlsx Excel file ...
Dependency:
    openpyxl
    xml.etree.ElementTree as etree
For more info see:
    - https://docs.python.org/2/library/xml.etree.elementtree.html
    - http://www.diveintopython3.net/xml.html
    - https://openpyxl.readthedocs.org/en/default/ """


import datetime
import os
import xml.etree.ElementTree

from types import StringType

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


class Myxml:
    
    def __init__(self, path="./", xmlfile=None):
        """
        Myxml class
        
        :param xmlfile: .xml file to be processed
        :type xmlfile: str
        """
        assert type(xmlfile) is StringType, "xmlfile is not a string: xmlfile is %r" % xmlfile
        # self.name = xmlfile
        self.absolute_path = os.path.join(path, xmlfile)
        self.workbook = Workbook()
        self.active_workbook = self.workbook.active

        # assert isinstance(xml.etree.ElementTree.parse(self.location).getroot, object)
        self.tree = xml.etree.ElementTree.parse(self.absolute_path)
        self.root = self.tree.getroot()
        self.first_row = None
        self.rows = None

    def __str__(self):
        pass

    def len_of_first_row_in_root(self):
        return len(self.get_first_row_in_root())

    def get_first_row_in_root(self):
        if self.root is not None:
            first_row = [elem.tag for elem in self.root[0]]
            # print(first_row)  # Uncomment for debug
            return first_row

    def get_rows(self):
        # rows = [val.text for i in range(0, len(self.root)) for val in self.root[i]]
        # row = []
        # for i in range(0, len(self.root)):
        #    for val in root[i]:
        #        row.append(val.text)
        
        # print(rows)    # Uncomment for debug
        # --- Print Elements for Row
        rows = []
        i = 0
        while i < len(self.root):
            row = [val.text for val in self.root[i]]
            i += 1
            rows.append(row)

        #print rows viene chiamato un botto di volte
        return rows


def create_xlsx_from_xlm(Myxml):
    """
    Main method: 
    :type Myxml: Myxml
    """
    
    # recupero path e filename per il salvataggio
    # lo chiamo nello stesso modo del file .xml
    path = Myxml.absolute_path[-4]  # senza l'estensione...
    extension = '.xlsx'

    # imposto il nome del foglio
    Myxml.active_workbook.title = Myxml.root.tag

    # creo l'header ..
    Myxml.active_workbook.append(Myxml.get_first_row_in_root())
    # creo le altre rows ..
    i = 0
    while i < len(Myxml.get_rows()):
        Myxml.active_workbook.append(Myxml.get_rows()[i]) # non va..
        i += 1


    Myxml.workbook.save(filename=path+extension)

def get_header_cell():
    pass

def set_cell_style(cell):
    """
Set Style for header's cell
    :param cell: an XLSX cell
    """
    cell.font = Font(name='Calibri',
                     size=16,
                     bold=True,
                     italic=False,
                     vertAlign='baseline',
                     underline='none',
                     strike=False,
                     color='FF000000')

    cell.fill = PatternFill(fill_type=None,
                            start_color='FFFFFFFF',
                            end_color='FF000000')

    cell.border = Border(left=Side(border_style=None,
                                   color='FF000000'),
                         right=Side(border_style=None,
                                    color='FF000000'),
                         top=Side(border_style=None,
                                  color='FF000000'),
                         bottom=Side(border_style='double',
                                     color='FF000000'),
                         diagonal=Side(border_style=None,
                                       color='FF000000'),
                         diagonal_direction=0,
                         outline=Side(border_style=None,
                                      color='FF000000'),
                         vertical=Side(border_style=None,
                                       color='FF000000'),
                         horizontal=Side(border_style=None,
                                         color='FF000000')
                         )
    cell.alignment = Alignment(horizontal="center",
                               vertical="center",
                               text_rotation=0,
                               wrap_text=True,
                               shrink_to_fit=True,
                               indent=0)
    cell.number_format = 'General'
    cell.protection = Protection(locked=True,
                                 hidden=False)

def set_header_styly(Workbook):
    pass

def main(name='book.xml'):
    """
    Main script (sostituito)
    """
    tree = xml.etree.ElementTree.parse(name)  # "importo" un file .xml 
    root = tree.getroot()  # ottengo la root del file, root è come una lista multidim
    
    # Create a Workbook per interfacciarmi al file Excel
    workbook = Workbook() # è una lista?
    
    # prendo qoello appena creato
    active_workbook = workbook.active
    active_workbook.title = root.tag
    # Creo l'header
    
    # root path where there are iterable object..
     
    user_root = root[0] 
    
    print root[0][1].attrib['name']
    print root[0][1]
    # and name to show in the header of excell Tablle ex: .attrib[name] 
    # will crash if .attrib[] doesn't exist ...
    header = [elem.attrib['name'] for elem in user_root] # [ root[0][i], .. ]  
    active_workbook.append(header)

    # set stiles for header: prendo la prima riga
    # active_workbook.iter_rows(row_offset=0) = <generator object get_squared_range at 0x01101B70>
    for row in active_workbook.iter_rows(row_offset=0): # row = (<Cell catalog.A1>, <Cell catalog.B1>, <Cell catalog.C1>, <Cell catalog.D1>, <Cell catalog.E1>, <Cell catalog.F1>)
        for cell in row:
            set_cell_style(cell)

    # Print elements for Row
    i = 1 #salto il primo elemento
    while i < len(user_root):
        row = [val.text for val in user_root[i]]
        active_workbook.append(row)
        i += 1
        
    # Save the file
    dest_filename = 'xmlTest2.xlsx'
    workbook.save(filename=dest_filename)


if __name__ == '__main__':
    print datetime.datetime.today()
    # --- New 
    a = Myxml(xmlfile='book.xml')
    create_xlsx_from_xlm(a)
    # --- Old
    main('./test1.xml')
    
