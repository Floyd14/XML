# -*- coding: UTF-8 -*-
""" xml_To_xlsx Test1.
This program convert .xml file to .xlsx Excel file ...
Dependency:
    openpyxl
    xml.etree.ElementTree as etree
For more info see:
    - https://docs.python.org/2/library/xml.etree.elementtree.html
    - http://www.diveintopython3.net/xml.html
    - https://openpyxl.readthedocs.org/en/default/ """

import xml.etree.ElementTree as ET
from openpyxl import Workbook


def main():
    """
    Main script
    """
    xml_filename = "./test1.xml"
    
    tree = ET.parse(xml_filename)           # "importo" un file .xml
    root = tree.getroot()                   # ottengo la root del file, root è come una lista multidim

    # trovo i manage object è una lista
    entry = root[0]
    managed_objects =  entry.findall('{raml20.xsd}managedObject')
    
    myObjs = []
    
    # per tutti gli oggetti
    for managed_object in managed_objects:
        myObj = [] # l'oggetto che voglio creare

        lac = ""
        lac_val = ""
        target = ""
        # per tutti i p
        for p in managed_object.findall('{raml20.xsd}p'):

            if p.attrib['name'] == "AdjgLAC":
                lac = p.attrib['name']
                lac_val = p.text
            if  p.attrib['name'] == "TargetCellDN":
                target = p.text[10:-10]

        name = managed_object.attrib['name']
        sorgente = name[:7]
        destinazione = name[11:]
        
        classe = managed_object.attrib['class']

        myObj = [sorgente, destinazione, classe, lac_val, target] # oggetto creato ...
        print 'Sorgente: %s, Destinazione: %s, Tipo: %s, Lac: %s, Lac_val: %s, Tar: %s' %(sorgente, destinazione, classe, lac, lac_val, target)
        myObjs.append(myObj)
        
    print "myObjs= ", myObjs

    # Create a Workbook per interfacciarmi al file Excel
    workbook = Workbook()
    my_workbook = workbook.active  # � una lista?
    
    # prendo qoello appena creato
    my_workbook.title = root.tag
    header = ['Sorgente',"Destinazione", "Classe", "Lac", "Targhet"]
    my_workbook.append(header)
    
    for obj in myObjs:
        my_workbook.append(obj)
    
    # Save the file
    dest_filename = 'xmlTest3.xlsx'
    workbook.save(filename=dest_filename)
    
    #def mg_get_class(managed_objects):
    #    [cls for managed_object. in managed_objects]
    #    return
       
#     # set stiles for header: prendo la prima riga
#     # active_workbook.iter_rows(row_offset=0) = <generator object get_squared_range at 0x01101B70>
#     for row in active_workbook.iter_rows(row_offset=0): # row = (<Cell catalog.A1>, <Cell catalog.B1>, <Cell catalog.C1>, <Cell catalog.D1>, <Cell catalog.E1>, <Cell catalog.F1>)
#         for cell in row:
#             set_cell_style(cell)

if __name__ == '__main__':
    main()

